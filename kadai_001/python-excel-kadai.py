import openpyxl
from datetime import datetime

workbook = openpyxl.Workbook()


adress1 = "株式会社ABC"
adress2 = "〒101-0022 東京都千代田区神田練塀町300"
adress3 = "TEL:03-1234-5678 FAX:03-1234-5678"
adress4 = "担当者名:鈴木一郎 様"


today = datetime.today()

ws = workbook.active

ws['A2'] = '請求書'
ws['A4'] = adress1
ws['A5'] = adress2
ws['A6'] = adress3
ws['A7'] = adress4


workbook.save('sample-1.xlsx')
workbook = openpyxl.load_workbook('sample-1.xlsx')

ws = workbook.active

ws['E4'] = 'No.'
ws['F4'] = '0001' 
ws['E5'] = '日付'
ws['F5'] = f"{today.strftime('%Y/%m/%d')}"

header = ['商品名', '数量', '単価', '金額']
ws.append(header)

data = [
    ['商品A', 2, 10000, 20000],
    ['商品B', 1, 15000, 15000],
    ['', '', '', 35000],
    ['', '', '', ''],
    ['合計', '', '', 35000],
    ['消費税', '', '', 3500],
    ['税込合計', '', '', 38500]
]


for row in data:
    ws.append(row)
ws.insert_cols(1,1)



workbook.save('売上表.xlsx')

